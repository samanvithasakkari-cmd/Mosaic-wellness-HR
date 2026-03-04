"""
Mosaic Talent — Full-Stack Flask Backend
SQLite persistence + Excel export + REST API
Deploy: python app.py  |  Production: gunicorn app:app
"""

import os, json, uuid, sqlite3, io
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template, g

app = Flask(__name__)
app.config['DATABASE'] = os.path.join(os.path.dirname(__file__), 'instance', 'mosaic.db')

# ─── Database helpers ─────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        os.makedirs(os.path.dirname(app.config['DATABASE']), exist_ok=True)
        g.db = sqlite3.connect(
            app.config['DATABASE'],
            detect_types=sqlite3.PARSE_DECLTYPES
        )
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS jobs (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            parsed_data TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS assessments (
            id TEXT PRIMARY KEY,
            job_id TEXT NOT NULL,
            questions TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(job_id) REFERENCES jobs(id)
        );

        CREATE TABLE IF NOT EXISTS candidates (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            assessment_id TEXT NOT NULL,
            status TEXT DEFAULT 'registered',
            score REAL,
            evaluation TEXT,
            answers TEXT,
            ai_probability REAL,
            fairness_score REAL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(assessment_id) REFERENCES assessments(id)
        );

        CREATE UNIQUE INDEX IF NOT EXISTS idx_candidate_email_assessment
            ON candidates(email, assessment_id);

        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_type TEXT NOT NULL,
            user_email TEXT,
            details TEXT,
            ip_address TEXT,
            timestamp TEXT DEFAULT (datetime('now'))
        );
    """)
    db.commit()

    # Seed demo data if empty
    row = db.execute("SELECT count(*) as c FROM jobs").fetchone()
    if row['c'] == 0:
        _seed_demo(db)

def _seed_demo(db):
    jid = 'demo-job-1'
    aid = 'demo-assess-1'
    parsed = {
        "roleTitle": "Senior Product Designer",
        "seniority": "Senior",
        "department": "Design",
        "domain": "Fintech",
        "isHybrid": False,
        "hybridDepts": [],
        "bloomLevels": ["Analyze", "Evaluate", "Create"],
        "difficultyLevel": "Advanced",
        "keySkills": [
            {"skill": "UI Design", "bloomLevel": "Create"},
            {"skill": "User Research", "bloomLevel": "Analyze"},
            {"skill": "Design Systems", "bloomLevel": "Evaluate"}
        ],
        "yearsOfExperience": "5+",
        "coreResponsibilities": ["Lead design projects", "Define design systems", "Mentor junior designers"]
    }
    questions = [
        {"id":"q1","type":"mcq","bloomLevel":"Remember","text":"What is the primary purpose of a design system?","options":["Consistency & scalability","Developer speed","Marketing alignment","Cost reduction"],"correctAnswer":"Consistency & scalability","competency":"Design Systems","difficulty":"Foundational"},
        {"id":"q2","type":"mcq","bloomLevel":"Understand","text":"Which principle is central to User-Centered Design?","options":["Iterative design process","System-first thinking","Developer convenience","Marketing priority"],"correctAnswer":"Iterative design process","competency":"UX Principles","difficulty":"Foundational"},
        {"id":"q3","type":"mcq","bloomLevel":"Apply","text":"When would you use a card-sorting exercise?","options":["To understand users' mental models","To test visual aesthetics","To measure load time","To create wireframes"],"correctAnswer":"To understand users' mental models","competency":"User Research","difficulty":"Applied"},
        {"id":"q4","type":"mcq","bloomLevel":"Analyze","text":"A user funnel shows 60% drop-off at checkout. Which analysis method is MOST appropriate first?","options":["Session recording + heatmap analysis","Full redesign","A/B test immediately","Stakeholder meeting"],"correctAnswer":"Session recording + heatmap analysis","competency":"Problem Solving","difficulty":"Advanced"},
        {"id":"q5","type":"mcq","bloomLevel":"Evaluate","text":"You have two design directions from user testing. One scores higher on desirability, one on usability. How do you decide?","options":["Weigh against business goals and task completion rates","Pick the more beautiful one","Ask engineers which is easier","Go with stakeholder preference"],"correctAnswer":"Weigh against business goals and task completion rates","competency":"Decision Making","difficulty":"Advanced"},
        {"id":"q6","type":"short","bloomLevel":"Understand","text":"Explain the difference between UX and UI design with a real-world example.","competency":"Design Fundamentals","difficulty":"Foundational"},
        {"id":"q7","type":"short","bloomLevel":"Apply","text":"Walk through your process for conducting a usability test from planning to insights.","competency":"User Research","difficulty":"Applied"},
        {"id":"q8","type":"short","bloomLevel":"Analyze","text":"How do you measure the success of a design change post-launch? List specific metrics.","competency":"Data-Driven Design","difficulty":"Advanced"},
        {"id":"q9","type":"scenario","bloomLevel":"Evaluate","text":"A developer tells you a key design component is too complex to build in the current sprint. The PM wants it shipped. How do you navigate this?","competency":"Collaboration","difficulty":"Advanced"},
        {"id":"q10","type":"scenario","bloomLevel":"Evaluate","text":"Stakeholder A wants the CTA button red. Stakeholder B wants it green. User data shows no significant difference. What do you do?","competency":"Stakeholder Management","difficulty":"Advanced"},
        {"id":"q11","type":"case_study","bloomLevel":"Create","text":"Design a complete strategy to redesign a legacy enterprise dashboard with 100+ features. It currently has a 40% task-completion rate. Outline your discovery, prioritization, rollout, and success metrics.","competency":"Strategic Reasoning","difficulty":"Expert"}
    ]
    db.execute("INSERT INTO jobs (id,title,description,parsed_data) VALUES (?,?,?,?)",
               (jid, "Senior Product Designer", "Demo job description", json.dumps(parsed)))
    db.execute("INSERT INTO assessments (id,job_id,questions) VALUES (?,?,?)",
               (aid, jid, json.dumps(questions)))

    candidates_data = [
        ("c1","Alex Rivera","alex@example.com", 84, 0.12, 0.95,
         '{"scores":{"accuracy":4,"relevance":5,"reasoning":4,"competency":4,"culture":4},"totalScore":84,"reasoning":"Alex demonstrated strong understanding of design principles. Responses showed depth and practical fintech experience across all Bloom levels.","strengths":["Visual Design","Strategic Thinking","User Empathy","Stakeholder Communication"],"weaknesses":["Technical Prototyping Depth","Quantitative Metrics"],"recommendation":"Advance","aiProbability":0.12,"fairnessConfidence":0.95}',
         '{"q1":"Consistency & scalability","q2":"Iterative design process","q3":"To understand users mental models","q4":"Session recording + heatmap analysis","q5":"Weigh against business goals and task completion rates","q6":"UX is the overall journey and feel of a product — e.g. how easy Netflix is to navigate. UI is the visual presentation — the button colors, typography, layout. UX without UI is a skeleton; UI without UX is decoration.","q7":"I start by defining goals and success criteria, recruit 5-8 representative users, script tasks without leading questions, run moderated sessions, then synthesize findings with affinity mapping.","q8":"I track task completion rate, time-on-task, error rate, SUS score post-launch, and retention delta in the redesigned flow.","q9":"I would first understand the technical constraint in detail, then explore phased delivery — ship a simplified version now, log it as design debt, and schedule the full component in the next sprint with a tech spec.","q10":"I would reframe it: this is not a color debate. I would propose a 2-week A/B test with a clear success metric (CTR + conversion), present findings to both stakeholders, and let data decide.","q11":"Phase 1: Discovery — audit all 100+ features against usage analytics, identify the 20% used by 80% of users. Phase 2: Jobs-to-be-Done research with 12 power users. Phase 3: Prioritize by impact vs complexity matrix. Phase 4: Incremental rollout with feature flags, measuring task completion lift. Success: 40% → 70% task completion in 6 months."}'),
        ("c2","Jordan Kim","jordan@example.com", 71, 0.08, 0.91,
         '{"scores":{"accuracy":3,"relevance":4,"reasoning":4,"competency":3,"culture":4},"totalScore":71,"reasoning":"Jordan showed solid foundational knowledge and good collaboration instincts but lacked depth in strategic and data-driven responses.","strengths":["Collaboration","Communication","Empathy"],"weaknesses":["Advanced Prototyping","Metrics Definition","System Design"],"recommendation":"Review","aiProbability":0.08,"fairnessConfidence":0.91}',
         '{"q1":"Consistency & scalability","q2":"Iterative design process","q3":"To understand users mental models","q4":"Session recording + heatmap analysis","q5":"Weigh against business goals and task completion rates","q6":"UX is about the experience, UI is about looks.","q7":"Recruit users, give them tasks, watch them, take notes.","q8":"I look at user feedback and analytics.","q9":"I would talk to both the developer and PM to find a middle ground.","q10":"I would suggest testing both colors with users.","q11":"I would start by talking to users about what they need and then redesign based on that feedback."}'),
        ("c3","Sam Patel","sam@example.com", 55, 0.67, 0.72,
         '{"scores":{"accuracy":3,"relevance":3,"reasoning":2,"competency":3,"culture":2},"totalScore":55,"reasoning":"Responses were generic and formulaic. High AI probability detected — answers lacked personal experience markers and used templated language patterns.","strengths":["Basic Design Knowledge"],"weaknesses":["Strategic Thinking","Depth of Experience","Critical Reasoning","Originality"],"recommendation":"Reject","aiProbability":0.67,"fairnessConfidence":0.72}',
         '{"q1":"Consistency & scalability","q2":"Iterative design process","q3":"To understand users mental models","q4":"Session recording + heatmap analysis","q5":"Weigh against business goals and task completion rates","q6":"UX design focuses on the overall user experience while UI design deals with the visual elements.","q7":"User testing involves planning, recruiting, conducting sessions and analyzing results.","q8":"Success can be measured through various KPIs and metrics.","q9":"Communication and collaboration are key to resolving such conflicts.","q10":"Data-driven decision making should be used to resolve design disagreements.","q11":"A comprehensive redesign strategy would involve research, planning, execution and measurement phases."}'),
    ]
    for cid, name, email, score, ai_prob, fair, eval_json, ans_json in candidates_data:
        db.execute("""INSERT INTO candidates (id,name,email,assessment_id,status,score,evaluation,answers,ai_probability,fairness_score)
                      VALUES (?,?,?,?,?,?,?,?,?,?)""",
                   (cid, name, email, aid, 'completed', score, eval_json, ans_json, ai_prob, fair))

    logs = [
        ("TEST_SUBMITTED", "alex@example.com", "Score: 84% | Assessment: demo-assess-1"),
        ("TEST_SUBMITTED", "jordan@example.com", "Score: 71% | Assessment: demo-assess-1"),
        ("CANDIDATE_REGISTERED", "sam@example.com", "Registered for assessment demo-assess-1"),
        ("ASSESSMENT_PUBLISHED", "recruiter@mosaic.ai", "Published: Senior Product Designer (demo-assess-1)"),
    ]
    for evt, email, details in logs:
        db.execute("INSERT INTO logs (event_type,user_email,details) VALUES (?,?,?)", (evt, email, details))
    db.commit()

# ─── Request context ──────────────────────────────────────────────────────────
@app.before_request
def setup():
    with app.app_context():
        pass

# ─── Serve frontend ───────────────────────────────────────────────────────────
@app.route('/')
@app.route('/test')
def index():
    with app.app_context():
        init_db()
    return render_template('index.html')

# ─── API: Stats ───────────────────────────────────────────────────────────────
@app.route('/api/stats')
def api_stats():
    db = get_db()
    active = db.execute("SELECT count(*) as c FROM assessments").fetchone()['c']
    total  = db.execute("SELECT count(*) as c FROM candidates").fetchone()['c']
    avg_row = db.execute("SELECT AVG(score) as a FROM candidates WHERE score IS NOT NULL").fetchone()
    avg    = round(avg_row['a'] or 0)
    advance = db.execute("SELECT count(*) as c FROM candidates WHERE json_extract(evaluation,'$.recommendation')='Advance'").fetchone()['c']
    return jsonify({"active": active, "candidates": total, "avgScore": avg, "advanceCount": advance})

# ─── API: Jobs ────────────────────────────────────────────────────────────────
@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    db = get_db()
    rows = db.execute("SELECT * FROM jobs ORDER BY created_at DESC").fetchall()
    return jsonify([{**dict(r), 'parsed_data': json.loads(r['parsed_data'] or '{}')} for r in rows])

@app.route('/api/jobs', methods=['POST'])
def create_job():
    data = request.json
    db = get_db()
    db.execute("INSERT INTO jobs (id,title,description,parsed_data) VALUES (?,?,?,?)",
               (data['id'], data['title'], data['description'], json.dumps(data['parsed_data'])))
    db.commit()
    db.execute("INSERT INTO logs (event_type,user_email,details) VALUES (?,?,?)",
               ('JOB_CREATED', 'recruiter', f"Created job: {data['title']}"))
    db.commit()
    return jsonify({"success": True})

# ─── API: Assessments ─────────────────────────────────────────────────────────
@app.route('/api/assessments', methods=['POST'])
def create_assessment():
    data = request.json
    db = get_db()
    db.execute("INSERT INTO assessments (id,job_id,questions) VALUES (?,?,?)",
               (data['id'], data['job_id'], json.dumps(data['questions'])))
    db.commit()
    db.execute("INSERT INTO logs (event_type,user_email,details) VALUES (?,?,?)",
               ('ASSESSMENT_PUBLISHED', 'recruiter', f"Published assessment {data['id']} for job {data['job_id']}"))
    db.commit()
    return jsonify({"success": True})

@app.route('/api/assessments/<aid>')
def get_assessment(aid):
    db = get_db()
    row = db.execute("SELECT * FROM assessments WHERE id=?", (aid,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify({**dict(row), 'questions': json.loads(row['questions'])})

# ─── API: Candidates ──────────────────────────────────────────────────────────
@app.route('/api/candidates', methods=['GET'])
def get_candidates():
    db = get_db()
    rows = db.execute("""
        SELECT c.*, j.title as job_title, a.questions
        FROM candidates c
        JOIN assessments a ON c.assessment_id = a.id
        JOIN jobs j ON a.job_id = j.id
        ORDER BY c.score DESC NULLS LAST, c.created_at DESC
    """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d['evaluation'] = json.loads(d['evaluation']) if d['evaluation'] else None
        d['answers'] = json.loads(d['answers']) if d['answers'] else {}
        d['questions'] = json.loads(d['questions']) if d['questions'] else []
        result.append(d)
    return jsonify(result)

@app.route('/api/candidates/<cid>')
def get_candidate(cid):
    db = get_db()
    row = db.execute("""
        SELECT c.*, j.title as job_title, a.questions, j.parsed_data as job_parsed
        FROM candidates c
        JOIN assessments a ON c.assessment_id = a.id
        JOIN jobs j ON a.job_id = j.id
        WHERE c.id=?
    """, (cid,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    d = dict(row)
    d['evaluation'] = json.loads(d['evaluation']) if d['evaluation'] else None
    d['answers'] = json.loads(d['answers']) if d['answers'] else {}
    d['questions'] = json.loads(d['questions']) if d['questions'] else []
    d['job_parsed'] = json.loads(d['job_parsed']) if d['job_parsed'] else {}
    return jsonify(d)

@app.route('/api/candidates', methods=['POST'])
def create_candidate():
    data = request.json
    db = get_db()
    try:
        # Check if exists
        existing = db.execute(
            "SELECT id FROM candidates WHERE email=? AND assessment_id=?",
            (data['email'], data['assessment_id'])
        ).fetchone()
        if existing:
            return jsonify({"success": True, "id": existing['id']})
        
        cid = data.get('id', str(uuid.uuid4())[:8])
        db.execute(
            "INSERT INTO candidates (id,name,email,assessment_id,status) VALUES (?,?,?,?,?)",
            (cid, data['name'], data['email'], data['assessment_id'], 'registered')
        )
        db.execute("INSERT INTO logs (event_type,user_email,details,ip_address) VALUES (?,?,?,?)",
                   ('CANDIDATE_REGISTERED', data['email'],
                    f"Registered for {data['assessment_id']}",
                    request.remote_addr))
        db.commit()
        return jsonify({"success": True, "id": cid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/candidates/<cid>', methods=['PUT'])
def update_candidate(cid):
    data = request.json
    db = get_db()
    db.execute("""
        UPDATE candidates SET
            status=?, score=?, evaluation=?, answers=?,
            ai_probability=?, fairness_score=?
        WHERE id=?
    """, (
        data.get('status', 'completed'),
        data.get('score'),
        json.dumps(data.get('evaluation')) if data.get('evaluation') else None,
        json.dumps(data.get('answers', {})),
        data.get('ai_probability'),
        data.get('fairness_score'),
        cid
    ))
    db.execute("INSERT INTO logs (event_type,user_email,details,ip_address) VALUES (?,?,?,?)",
               ('TEST_SUBMITTED', data.get('email', 'unknown'),
                f"Score: {data.get('score')}% | Assessment: {data.get('assessment_id','')}",
                request.remote_addr))
    db.commit()
    return jsonify({"success": True})

@app.route('/api/candidates/<cid>', methods=['DELETE'])
def delete_candidate(cid):
    db = get_db()
    db.execute("DELETE FROM candidates WHERE id=?", (cid,))
    db.commit()
    return jsonify({"success": True})

# ─── API: Logs ────────────────────────────────────────────────────────────────
@app.route('/api/logs')
def get_logs():
    db = get_db()
    rows = db.execute("SELECT * FROM logs ORDER BY timestamp DESC LIMIT 200").fetchall()
    return jsonify([dict(r) for r in rows])

# ─── API: Excel Export ────────────────────────────────────────────────────────
@app.route('/api/export/excel')
def export_excel():
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    db = get_db()
    rows = db.execute("""
        SELECT c.*, j.title as job_title, a.questions
        FROM candidates c
        JOIN assessments a ON c.assessment_id = a.id
        JOIN jobs j ON a.job_id = j.id
        ORDER BY c.score DESC NULLS LAST
    """).fetchall()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Leaderboard"

    # Styles
    hdr_fill = PatternFill("solid", fgColor="0F172A")   # slate-950
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    emerald_fill = PatternFill("solid", fgColor="10B981")
    amber_fill   = PatternFill("solid", fgColor="F59E0B")
    red_fill     = PatternFill("solid", fgColor="EF4444")
    gold_fill    = PatternFill("solid", fgColor="FCD34D")
    silver_fill  = PatternFill("solid", fgColor="CBD5E1")
    bronze_fill  = PatternFill("solid", fgColor="FED7AA")
    alt_fill     = PatternFill("solid", fgColor="F8FAFC")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(border_style="thin", color="E2E8F0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws1.merge_cells("A1:J1")
    title_cell = ws1["A1"]
    title_cell.value = "MOSAIC TALENT — CANDIDATE ASSESSMENT RESULTS"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor="0F172A")
    title_cell.alignment = center
    ws1.row_dimensions[1].height = 36

    # Sub-title
    ws1.merge_cells("A2:J2")
    sub = ws1["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}  |  Powered by Gemini AI"
    sub.font = Font(size=10, color="64748B", name="Calibri")
    sub.fill = PatternFill("solid", fgColor="F8FAFC")
    sub.alignment = center
    ws1.row_dimensions[2].height = 22

    # Headers
    headers = ["Rank", "Candidate Name", "Email", "Role", "Score", "Recommendation",
               "AI Usage %", "Fairness %", "Strengths", "Areas for Improvement"]
    ws1.row_dimensions[3].height = 28
    for ci, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    # Data rows
    for ri, row in enumerate(rows, 1):
        data_row = ri + 3
        ws1.row_dimensions[data_row].height = 48
        ev = json.loads(row['evaluation']) if row['evaluation'] else {}
        strengths = "; ".join(ev.get('strengths', []))
        weaknesses = "; ".join(ev.get('weaknesses', []))
        rec = ev.get('recommendation', 'Pending')
        score = row['score'] or 0
        ai_prob = (row['ai_probability'] or 0) * 100
        fair = (row['fairness_score'] or 0) * 100

        values = [ri, row['name'], row['email'], row['job_title'] or '',
                  f"{score}%", rec, f"{ai_prob:.0f}%", f"{fair:.0f}%",
                  strengths, weaknesses]

        row_fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        for ci, val in enumerate(values, 1):
            cell = ws1.cell(row=data_row, column=ci, value=val)
            cell.border = border
            cell.fill = row_fill

            # Rank medal colors
            if ci == 1:
                cell.alignment = center
                cell.font = Font(bold=True, size=13, name="Calibri")
                if ri == 1: cell.fill = gold_fill
                elif ri == 2: cell.fill = silver_fill
                elif ri == 3: cell.fill = bronze_fill
            elif ci == 5:  # Score
                cell.alignment = center
                cell.font = Font(bold=True, name="Calibri",
                                  color="047857" if score >= 75 else ("B45309" if score >= 50 else "B91C1C"))
            elif ci == 6:  # Recommendation
                cell.alignment = center
                cell.font = Font(bold=True, size=10, name="Calibri",
                                  color="FFFFFF" if rec != 'Pending' else "64748B")
                if rec == 'Advance':   cell.fill = emerald_fill
                elif rec == 'Review':  cell.fill = amber_fill
                elif rec == 'Reject':  cell.fill = red_fill
            elif ci in (7, 8):
                cell.alignment = center
            else:
                cell.alignment = left

    # Column widths
    col_widths = [8, 22, 28, 28, 10, 16, 12, 12, 45, 45]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Detailed Responses ──
    ws2 = wb.create_sheet("Candidate Responses")
    ws2.row_dimensions[1].height = 30
    ws2.row_dimensions[2].height = 22

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "CANDIDATE RESPONSES — DETAILED VIEW"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws2["A1"].fill = PatternFill("solid", fgColor="065F46")
    ws2["A1"].alignment = center

    ws2.merge_cells("A2:F2")
    ws2["A2"].value = "All candidate answers to individual questions"
    ws2["A2"].font = Font(size=10, color="64748B", name="Calibri")
    ws2["A2"].alignment = center

    resp_headers = ["Candidate", "Email", "Role", "Q#", "Question (Bloom Level)", "Answer"]
    ws2.row_dimensions[3].height = 28
    for ci, h in enumerate(resp_headers, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="065F46")
        cell.alignment = center
        cell.border = border

    resp_row = 4
    for row in rows:
        questions = json.loads(row['questions']) if row['questions'] else []
        answers   = json.loads(row['answers'])   if row['answers']   else {}
        ev        = json.loads(row['evaluation'])if row['evaluation']else {}
        cand_fill = PatternFill("solid", fgColor="ECFDF5")

        for qi, q in enumerate(questions, 1):
            ws2.row_dimensions[resp_row].height = max(60, len(answers.get(q['id'], '')) // 3)
            bloom = q.get('bloomLevel', '')
            q_text = f"[{bloom}] {q['text']}"
            ans_text = answers.get(q['id'], '(No answer provided)')

            vals = [row['name'] if qi == 1 else '', row['email'] if qi == 1 else '',
                    (row['job_title'] or '') if qi == 1 else '',
                    f"Q{qi}", q_text, ans_text]
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=resp_row, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if ci <= 3:
                    cell.fill = cand_fill
                    cell.font = Font(bold=(ci == 1), name="Calibri")
            resp_row += 1

        # Spacer row
        ws2.row_dimensions[resp_row].height = 8
        resp_row += 1

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 6
    ws2.column_dimensions['E'].width = 55
    ws2.column_dimensions['F'].width = 70

    # ── Sheet 3: Score Analytics ──
    ws3 = wb.create_sheet("Analytics")
    ws3.merge_cells("A1:D1")
    ws3["A1"].value = "SCORE ANALYTICS"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws3["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws3["A1"].alignment = center
    ws3.row_dimensions[1].height = 30

    analytics_data = [
        ("", ""),
        ("SUMMARY STATISTICS", ""),
        ("Total Candidates", len(rows)),
        ("Avg Score", f"{round(sum(r['score'] or 0 for r in rows)/max(len(rows),1))}%"),
        ("Top Score", f"{max((r['score'] or 0 for r in rows), default=0)}%"),
        ("Min Score", f"{min((r['score'] or 0 for r in rows), default=0)}%"),
        ("Advance Count", sum(1 for r in rows if json.loads(r['evaluation'] or '{}').get('recommendation')=='Advance')),
        ("Review Count",  sum(1 for r in rows if json.loads(r['evaluation'] or '{}').get('recommendation')=='Review')),
        ("Reject Count",  sum(1 for r in rows if json.loads(r['evaluation'] or '{}').get('recommendation')=='Reject')),
        ("High AI Prob (>50%)", sum(1 for r in rows if (r['ai_probability'] or 0) > 0.5)),
        ("", ""),
        ("SCORE DISTRIBUTION", ""),
        ("90-100%", sum(1 for r in rows if 90 <= (r['score'] or 0) <= 100)),
        ("75-89%",  sum(1 for r in rows if 75 <= (r['score'] or 0) < 90)),
        ("60-74%",  sum(1 for r in rows if 60 <= (r['score'] or 0) < 75)),
        ("40-59%",  sum(1 for r in rows if 40 <= (r['score'] or 0) < 60)),
        ("0-39%",   sum(1 for r in rows if 0  <= (r['score'] or 0) < 40)),
    ]
    for ri, (label, val) in enumerate(analytics_data, 2):
        if label in ("SUMMARY STATISTICS", "SCORE DISTRIBUTION"):
            cell = ws3.cell(row=ri, column=1, value=label)
            cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
            ws3.merge_cells(f"A{ri}:B{ri}")
        else:
            ws3.cell(row=ri, column=1, value=label).font = Font(name="Calibri")
            ws3.cell(row=ri, column=2, value=val).font = Font(bold=True, name="Calibri")

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 18

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"mosaic_talent_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ─── Run ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import os
    os.makedirs('instance', exist_ok=True)
    # Initialize DB before first request
    with app.app_context():
        init_db()
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    print(f"\n🚀 Mosaic Talent running → http://localhost:{port}\n")
    app.run(host='0.0.0.0', port=port, debug=debug)
